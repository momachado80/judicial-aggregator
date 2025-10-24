from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_
from typing import Optional
from io import BytesIO
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from ...database import get_db
from ...normalization.models import Process

router = APIRouter(prefix="/export", tags=["export"])


def format_currency(value: Optional[float]) -> str:
    """Formata valor monetário"""
    if value is None or value == 0:
        return "R$ 0,00"
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


@router.get("/pdf")
async def export_pdf(
    tribunal: Optional[str] = Query(None),
    classe_tpu: Optional[str] = Query(None),
    relevance: Optional[str] = Query(None),
    numero_cnj: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Exporta processos filtrados para PDF
    """
    # Construir query com filtros
    query = db.query(Process)
    
    filters = []
    if tribunal:
        filters.append(Process.tribunal == tribunal)
    if classe_tpu:
        filters.append(Process.classe_tpu == classe_tpu)
    if relevance:
        filters.append(Process.relevance == relevance)
    if numero_cnj:
        filters.append(Process.numero_cnj.ilike(f"%{numero_cnj}%"))
    
    if filters:
        query = query.filter(and_(*filters))
    
    # Buscar processos
    processes = query.order_by(Process.relevance.desc(), Process.valor_causa.desc()).all()
    
    # Criar PDF em memória
    buffer = BytesIO()
    
    # Configurar documento em paisagem (landscape) para caber mais colunas
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=1.5*cm
    )
    
    # Elementos do PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilo customizado para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#6366f1'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Estilo para subtítulo
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=10,
        alignment=TA_CENTER
    )
    
    # Título
    title = Paragraph("📊 Relatório de Processos Judiciais", title_style)
    elements.append(title)
    
    # Data e hora de geração
    now = datetime.now().strftime("%d/%m/%Y às %H:%M")
    subtitle = Paragraph(f"Gerado em {now} | Total: {len(processes)} processos", subtitle_style)
    elements.append(subtitle)
    
    # Filtros aplicados
    if any([tribunal, classe_tpu, relevance, numero_cnj]):
        filtros_text = "Filtros aplicados: "
        filtros_list = []
        if tribunal:
            filtros_list.append(f"Tribunal: {tribunal}")
        if classe_tpu:
            tipo = "Divórcio" if classe_tpu == "8015" else "Inventário"
            filtros_list.append(f"Tipo: {tipo}")
        if relevance:
            filtros_list.append(f"Relevância: {relevance}")
        if numero_cnj:
            filtros_list.append(f"CNJ: {numero_cnj}")
        
        filtros_text += " | ".join(filtros_list)
        filtros_para = Paragraph(filtros_text, subtitle_style)
        elements.append(filtros_para)
    
    elements.append(Spacer(1, 0.5*cm))
    
    # Preparar dados da tabela
    data = [
        ["CNJ", "Tribunal", "Tipo", "Comarca", "Vara", "Valor da Causa", "Relevância"]
    ]
    
    for proc in processes:
        tipo = "Divórcio" if proc.classe_tpu == "8015" else "Inventário"
        
        data.append([
            proc.numero_cnj or "N/A",
            proc.tribunal or "N/A",
            tipo,
            proc.comarca or "N/A",
            proc.vara or "N/A",
            format_currency(proc.valor_causa),
            proc.relevance or "N/A"
        ])
    
    # Criar tabela
    table = Table(data, colWidths=[5*cm, 2*cm, 2.5*cm, 3*cm, 3*cm, 3*cm, 2*cm])
    
    # Estilo da tabela
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Body
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),  # Valor da Causa alinhado à direita
        ('ALIGN', (6, 1), (6, -1), 'CENTER'),  # Relevância centralizada
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    
    # Adicionar rodapé com total
    elements.append(Spacer(1, 0.5*cm))
    
    total_valor = sum(p.valor_causa or 0 for p in processes)
    footer_text = f"<b>Total de processos:</b> {len(processes)} | <b>Valor total em causa:</b> {format_currency(total_valor)}"
    footer = Paragraph(footer_text, subtitle_style)
    elements.append(footer)
    
    # Construir PDF
    doc.build(elements)
    
    # Retornar PDF
    buffer.seek(0)
    
    filename = f"processos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/excel")
async def export_excel(
    tribunal: Optional[str] = Query(None),
    classe_tpu: Optional[str] = Query(None),
    relevance: Optional[str] = Query(None),
    numero_cnj: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Exporta processos filtrados para Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Construir query com filtros
    query = db.query(Process)
    
    filters = []
    if tribunal:
        filters.append(Process.tribunal == tribunal)
    if classe_tpu:
        filters.append(Process.classe_tpu == classe_tpu)
    if relevance:
        filters.append(Process.relevance == relevance)
    if numero_cnj:
        filters.append(Process.numero_cnj.ilike(f"%{numero_cnj}%"))
    
    if filters:
        query = query.filter(and_(*filters))
    
    # Buscar processos
    processes = query.order_by(Process.relevance.desc(), Process.valor_causa.desc()).all()
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Processos Judiciais"
    
    # Estilos
    header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Headers
    headers = [
        "Número CNJ",
        "Tribunal",
        "Tipo",
        "Comarca",
        "Vara",
        "Valor da Causa",
        "Relevância",
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Dados
    for row_num, proc in enumerate(processes, 2):
        tipo = "Divórcio" if proc.classe_tpu == "8015" else "Inventário"
        
        data = [
            proc.numero_cnj or "N/A",
            proc.tribunal or "N/A",
            tipo,
            proc.comarca or "N/A",
            proc.vara or "N/A",
            proc.valor_causa or 0,
            proc.relevance or "N/A",

        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border_style
            
            # Alinhamento
            if col_num == 6:  # Valor
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = 'R$ #,##0.00'
            elif col_num == 7:  # Relevância
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")
            
            # Cor de fundo alternada
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    # Ajustar largura das colunas
    column_widths = {
        'A': 30,  # CNJ
        'B': 12,  # Tribunal
        'C': 15,  # Tipo
        'D': 20,  # Comarca
        'E': 25,  # Vara
        'F': 18,  # Valor
        'G': 15,  # Relevância
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Congelar primeira linha
    ws.freeze_panes = 'A2'
    
    # Salvar em memória
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Nome do arquivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"processos_{timestamp}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
